#!/usr/bin/env python3
"""Generate the COMPLETE scCO2 system Bill of Materials as a formatted Excel file.

Covers: high-pressure fluid system, electronics/control, sensors, already-owned
equipment, and optional windowed vessel.

All links are direct product page URLs verified July 2026. Key safety corrections:
  - Tubing: must be 0.065" wall (0.035" wall rated only ~3,200 PSI — insufficient)
  - Relief valve: SS-4R3A (HIGH pressure, 6,000 PSI). SS-RL3S4 is LOW pressure (225 PSI max)
  - Relay: must accept 3.3V GPIO trigger (RPi-compatible optocoupler module)
  - Ball valves: only 1 needed (HiP valve on booster + SS-3NRM4 on vessel already cover other positions)
  - Pressure gauge: not needed (Duro United gauge already on vessel)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY  = "1E2761"
WHITE = "FFFFFF"


ROWS = [
    # ── 1. Motorized Needle Valve ─────────────────────────────────────────────
    dict(section="1. Motorized Needle Valve",
         item="Needle valve body", part="SS-1RS4",
         link_label="SS-1RS4",
         desc="Swagelok needle valve, regulating stem, 1/4\" tube compression",
         spec="316SS, 5,000 PSI @ 100°F, Cv = 0.37",
         vendor="Swagelok authorized distributor",
         qty=1, unit_price=300.00,
         link="https://products.swagelok.com/en/c/straight-pattern-regulating-stem/p/SS-1RS4?q=SS-1RS4"),
    dict(section="1. Motorized Needle Valve",
         item="Shaft coupler (NEMA 17 to valve stem)",
         part="5mm × 6.35mm flexible shaft coupler",
         link_label="Flexible Shaft Coupler",
         desc="Couples NEMA 17 stepper motor (5mm shaft) to SS-1RS4 valve stem (1/4\" = 6.35mm)",
         spec="Aluminum flexible coupling, 5mm × 6.35mm bores",
         vendor="Amazon",
         qty=1, unit_price=10.00,
         link="https://www.amazon.com/s?k=5mm+to+6.35mm+flexible+shaft+coupler"),

    # ── 2. Check Valve ────────────────────────────────────────────────────────
    dict(section="2. Check Valve",
         item="Check valve", part="SS-CHS4-5",
         link_label="SS-CHS4-5",
         desc="Swagelok check valve, 1/4\" tube fitting, 5 PSI cracking pressure",
         spec="316SS, 6,000 PSI (NOT SS-4C — that is only 3,000 PSI)",
         vendor="Swagelok authorized distributor",
         qty=1, unit_price=100.00,
         link="https://products.swagelok.com/en/c/fixed-pressure/p/SS-CHS4-5"),

    # ── 3. Manual Ball Valves ─────────────────────────────────────────────────
    dict(section="3. Manual Ball Valves",
         item="Ball valve — vessel inlet isolation", part="SS-83KS4",
         link_label="SS-83KS4",
         desc="Swagelok 83 series ball valve, 1/4\" tube fitting, PCTFE seats, CO2 compatible\n"
              "Only 1 needed — HiP valve on booster outlet covers supply shutoff position;\n"
              "SS-3NRM4 already on vessel outlet covers vent isolation position",
         spec="316SS, 6,000 PSI (NOT SS-43GS4 — that is only 3,000 PSI)",
         vendor="Swagelok authorized distributor",
         qty=1, unit_price=407.00,
         link="https://products.swagelok.com/en/c/2-way-straight-pattern/p/SS-83KS4"),

    # ── 4. Stainless Tubing ───────────────────────────────────────────────────
    dict(section="4. Stainless Tubing",
         item="SS tubing, 1/4\" OD × 0.065\" wall", part="5LVR1",
         link_label="5LVR1 Seamless SS Tubing",
         desc="Seamless 316SS tubing, 6 ft length — ASTM A213 + A269, annealed\n"
              "Buy 3 pieces (18 ft total) for adequate spare — system needs ~10 ft\n"
              "⚠ Must be seamless annealed (NOT welded) for Swagelok compression fittings\n"
              "⚠ Amazon FITOK B0BB1122VJ is only 0.035\" wall (~3,200 PSI) — NOT safe",
         spec="316SS seamless annealed, 1/4\" OD × 0.065\" wall, 6 ft per piece\n"
              "Rated 8,125 PSI @ 72°F — 2× safety margin over 4,061 PSI operating pressure",
         vendor="Grainger",
         qty=3, unit_price=36.43,
         link="https://www.grainger.com/search?searchQuery=5LVR1"),

    # ── 5. Compression Fittings ───────────────────────────────────────────────
    dict(section="5. Compression Fittings",
         item="Union (straight)", part="SS-400-6",
         link_label="SS-400-6",
         desc="Swagelok union fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Swagelok authorized distributor",
         qty=4, unit_price=15.00,
         link="https://products.swagelok.com/en/c/straights/p/SS-400-6?q=SS-400-6"),
    dict(section="5. Compression Fittings",
         item="90 degree elbow", part="SS-400-9",
         link_label="SS-400-9",
         desc="Swagelok 90 degree elbow fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Swagelok authorized distributor",
         qty=3, unit_price=20.00,
         link="https://products.swagelok.com/en/c/90-degree-elbows/p/SS-400-9?q=SS-400-9"),
    dict(section="5. Compression Fittings",
         item="Tee", part="SS-400-3",
         link_label="SS-400-3",
         desc="Swagelok tee fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Swagelok authorized distributor",
         qty=2, unit_price=28.00,
         link="https://products.swagelok.com/en/c/tees/p/SS-400-3?q=SS-400-3"),
    dict(section="5. Compression Fittings",
         item="End cap", part="SS-400-C",
         link_label="SS-400-C",
         desc="Swagelok end cap fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Swagelok authorized distributor",
         qty=3, unit_price=10.00,
         link="https://products.swagelok.com/en/c/caps/p/SS-400-C?q=SS-400-C"),

    # ── 6. Relief Valve & Vent ────────────────────────────────────────────────
    dict(section="6. Relief Valve",
         item="Relief valve", part="SS-4R3A",
         link_label="SS-4R3A",
         desc="Swagelok HIGH-PRESSURE proportional relief valve, 1/4\" tube compression\n"
              "⚠ MUST specify set pressure ~4,350 PSI (30 MPa) when ordering from distributor\n"
              "⚠ SS-RL3S4 (previously listed) is LOW-PRESSURE only (max 225 PSI) — DANGEROUS for this system",
         spec="316SS, 1/4\" tube, rated to 6,000 PSI\n"
              "Set pressure: 4,350 PSI (30 MPa) — between operating P (4,061 PSI) and MAWP (4,999 PSI)",
         vendor="Swagelok authorized distributor",
         qty=1, unit_price=250.00,
         link="https://products.swagelok.com/en/c/high-pressure-relief-valve/p/SS-4R3A?q=SS-4R3A"),
    dict(section="6. Relief Valve",
         item="Vent solenoid valve (automated depressurization)",
         part="Parker Series 34 HP / Asco 8290 HP",
         link_label="HP Solenoid Valve",
         desc="Normally-closed 24VDC solenoid valve, 1/4\" tube or NPT, HIGH-pressure rated\n"
              "Driven by RPi GPIO 18 via relay — opens automatically during DEPRESSURIZE state\n"
              "⚠ MUST be rated ≥ 6,000 PSI — standard solenoids (150-300 PSI) will fail catastrophically\n"
              "⚠ This part was MISSING from the original BOM",
         spec="316SS body, NC (fail-safe closed), 24VDC coil, 1/4\" process connection\n"
              "Candidates: Parker Series 34 HP, Asco 8290 HP, HiP solenoid valve",
         vendor="Parker / Asco / High Pressure Equipment",
         qty=1, unit_price=500.00,
         link="https://www.parker.com/us/en/search.html?q=high+pressure+solenoid+valve+24VDC+NC+6000+PSI"),
    dict(section="6. Relief Valve",
         item="PTFE thread tape", part="34P209",
         link_label="34P209 PTFE Tape",
         desc="PTFE sealing tape for NPT thread connections",
         spec="High-pressure / high-temp rated",
         vendor="Grainger",
         qty=1, unit_price=5.00,
         link="https://www.grainger.com/search?searchQuery=34P209"),

    # ── 7. Sensors ────────────────────────────────────────────────────────────
    dict(section="7. Sensors",
         item="Pressure transducer", part="5DEK9",
         link_label="5DEK9 Pressure Transducer",
         desc="Ashcroft G2, 0-5000 PSI, 4-20 mA output, nylon housing, 316SS wetted parts, IP67\n"
              "Grainger item 5DEK9 — Mfr model G17M0242F25000#\n"
              "⚠ Do NOT order K4708 — that is the 1-5V DC output version, incompatible with 4-20mA wiring",
         spec="4-20 mA → 1-5 V via 250 ohm shunt resistor into ADS1115 A0\n"
              "Nylon housing is fine for lab use; 316SS diaphragm is the CO2-wetted part",
         vendor="Grainger",
         qty=1, unit_price=200.00,
         link="https://www.grainger.com/product/ASHCROFT-Pressure-Transmitter-0-psi-5DEK9"),

    # ── 8. Electronics & Control ──────────────────────────────────────────────
    dict(section="8. Electronics & Control",
         item="Raspberry Pi 4 — CanaKit Starter Pro Kit (4 GB)", part="B07V5JTMV9",
         link_label="CanaKit RPi 4 Starter Kit",
         desc="CanaKit Starter Pro Kit: RPi4 4GB + case + 3.5A power supply + 32GB SD card + heatsinks\n"
              "SD card is INCLUDED in this kit — do NOT buy the separate microSD card",
         spec="4 GB RAM, includes 32GB pre-loaded SD card, case, power supply",
         vendor="Amazon — CanaKit",
         qty=1, unit_price=100.00,
         link="https://www.amazon.com/CanaKit-Raspberry-4GB-Starter-Kit/dp/B07V5JTMV9"),
    dict(section="8. Electronics & Control",
         item="ADS1115 16-bit I2C ADC module", part="B0DP43DDZG",
         link_label="ADS1115 ADC Module",
         desc="Reads pressure transducer (A0) and temperature sensor (A1)",
         spec="16-bit resolution, I2C interface, Raspberry Pi compatible",
         vendor="Amazon — Qoroos (3-pack)",
         qty=1, unit_price=10.00,
         link="https://www.amazon.com/Qoroos-Converter-Programmable-Amplifier-Development/dp/B0DP43DDZG"),
    dict(section="8. Electronics & Control",
         item="NEMA 17 stepper motor", part="B00PNEQKC0",
         link_label="NEMA 17 Stepper Motor",
         desc="Actuates the motorized needle valve via shaft coupler",
         spec="200 steps/rev, bipolar, NEMA 17 frame",
         vendor="Amazon — STEPPERONLINE",
         qty=1, unit_price=15.00,
         link="https://www.amazon.com/STEPPERONLINE-Stepper-Bipolar-Connector-compatible/dp/B00PNEQKC0"),
    dict(section="8. Electronics & Control",
         item="A4988 stepper driver", part="B07BND65C8",
         link_label="A4988 Stepper Driver",
         desc="Drives NEMA 17 at 1/16 microstepping; GPIO 17 STEP, 27 DIR, 22 EN",
         spec="1/16 microstepping driver module",
         vendor="Amazon — HiLetgo",
         qty=1, unit_price=8.00,
         link="https://www.amazon.com/HiLetgo-Stepstick-Stepper-Printer-Compatible/dp/B07BND65C8"),
    dict(section="8. Electronics & Control",
         item="Relay module — RPi-compatible (3.3V trigger)", part="B095YD3732",
         link_label="AEDIKO Relay Module",
         desc="AEDIKO 1-channel optocoupler-isolated relay module\n"
              "RPi GPIO 18 (3.3V) → relay IN → switches 24VDC to solenoid\n"
              "⚠ GAEYAELE B07DYLKH74 (prev BOM) requires 24VDC trigger — NOT compatible with RPi 3.3V GPIO",
         spec="5V coil, optocoupler isolated, accepts 3.3V RPi GPIO trigger, 1-channel",
         vendor="Amazon — AEDIKO",
         qty=1, unit_price=10.00,
         link="https://www.amazon.com/AEDIKO-Channel-Optocoupler-Isolation-Support/dp/B095YD3732"),
    dict(section="8. Electronics & Control",
         item="Jumper wires assorted kit", part="Dupont Jumper Wires",
         link_label="Jumper Wires Kit",
         desc="Male-to-female and male-to-male jumper wires for GPIO connections\n"
              "RPi GPIO → ADS1115 (4 wires), A4988 (3 wires), relay module (3 wires)",
         spec="120-piece assorted kit: male-to-male + male-to-female",
         vendor="Amazon",
         qty=1, unit_price=7.00,
         link="https://www.amazon.com/s?k=120+piece+jumper+wire+kit+male+female"),
    dict(section="8. Electronics & Control",
         item="DIN rail relay, 24VDC coil", part="4WAU2",
         link_label="4WAU2 DIN Rail Relay",
         desc="DIN-rail mounted relay for panel-style wiring of solenoid circuit (optional for bench setup)",
         spec="24VDC coil",
         vendor="Grainger",
         qty=1, unit_price=20.00,
         link="https://www.grainger.com/search?searchQuery=DIN+rail+relay+24VDC+coil"),
    dict(section="8. Electronics & Control",
         item="DIN rail power supply, 24VDC 50W", part="33NT20",
         link_label="33NT20 DIN Rail PSU",
         desc="Dayton DIN rail PSU — powers relay, solenoid, and sensor circuits",
         spec="24VDC, 50W output",
         vendor="Grainger",
         qty=1, unit_price=55.00,
         link="https://www.grainger.com/search?searchQuery=33NT20"),
]

ALREADY_OWNED = [
    dict(section="9. Already Owned",
         item="Pressure vessel", part="PARR 2302HC",
         desc="316SS pressure vessel, S/N 4540-1803-78845A, 1 L volume",
         spec="MAWP 5,000 PSI (34.47 MPa) @ 350°C — no purchase needed"),
    dict(section="9. Already Owned",
         item="Gas booster pump", part="HII 5G-TD-28/150-CO2",
         desc="Air-driven, twin-drive CO2 service booster, S/N 1205101",
         spec="Max outlet 25,000 PSIG (172 MPa) — no purchase needed"),
    dict(section="9. Already Owned",
         item="HiP valve on booster outlet", part="High Pressure Equipment valve",
         desc="Manual shutoff valve already installed on booster HP outlet — covers BV-1 position",
         spec="No purchase needed — saves ~$407 vs buying SS-83KS4 here"),
    dict(section="9. Already Owned",
         item="SS-3NRM4 needle valve on vessel outlet", part="Swagelok SS-3NRM4",
         desc="Manual needle valve already installed on vessel outlet — covers vent isolation position\n"
              "⚠ Must remain OPEN during automated operation — solenoid SV-1 controls automated venting",
         spec="8,000 PSI rated — no purchase needed — saves ~$407 vs buying SS-83KS4 here"),
    dict(section="9. Already Owned",
         item="Pressure gauge on vessel", part="Duro United 0-5000 PSI",
         desc="Mechanical pressure gauge already installed on PARR vessel port",
         spec="No purchase needed — saves ~$100 vs K4201"),
    dict(section="9. Already Owned",
         item="Temperature controller", part="INKBIRD",
         desc="60C setpoint temperature controller",
         spec="±0.01°C settled stability — no purchase needed"),
    dict(section="9. Already Owned",
         item="CO2 gas cylinder", part="UN1013 Bone Dry",
         desc="Primary process gas — Metro Welding Supply, Detroit MI",
         spec="No purchase needed (existing supply)"),
    dict(section="9. Already Owned",
         item="N2 gas cylinder", part="UN1066 Ultra High Purity",
         desc="Alternate experiment gas — Metro Welding Supply, Detroit MI",
         spec="No purchase needed (existing supply)"),
    dict(section="9. Already Owned",
         item="Ar gas cylinder", part="UN1006 Prepurified",
         desc="Alternate experiment gas — Metro Welding Supply, Detroit MI",
         spec="No purchase needed (existing supply)"),
    dict(section="9. Already Owned",
         item="Air gas cylinder (drive air)", part="UN1002 Dry Grade",
         desc="Drives booster pump only — never enters vessel",
         spec="No purchase needed (existing supply)"),
]

OPTIONAL = [
    dict(section="10. Optional: Windowed Vessel",
         item="Windowed pressure vessel (sapphire window)", part="EZE-Seal series",
         link_label="EZE-Seal Windowed Reactor",
         desc="Autoclave Engineers (Parker) windowed reactor for flow visualization",
         spec="Sapphire window, up to 60 MPa",
         vendor="Autoclave Engineers (Parker)",
         qty=1, unit_price=5000.00,
         link="https://www.autoclaveengineers.com/search/?q=EZE-Seal+windowed+reactor"),
    dict(section="10. Optional: Windowed Vessel",
         item="Windowed pressure vessel (alternate)", part="Sitec 100-600 mL",
         link_label="Sitec Windowed Reactor",
         desc="Sitec Reactor Technology windowed reactor, sapphire/borosilicate",
         spec="Up to 100 MPa, 100-600 mL volume",
         vendor="Sitec Reactor Technology",
         qty=1, unit_price=4500.00,
         link="https://www.sitec-ag.ch/?s=windowed+pressure+reactor+sapphire"),
]

SECTION_COLORS = {
    "1. Motorized Needle Valve": "E0F4F7",
    "2. Check Valve": "FFF3E0",
    "3. Manual Ball Valves": "E8EAF6",
    "4. Stainless Tubing": "F1F8E9",
    "5. Compression Fittings": "FCE4EC",
    "6. Relief Valve": "FFEBEE",
    "7. Sensors": "E3F2FD",
    "8. Electronics & Control": "FFF9C4",
    "9. Already Owned": "ECEFF1",
    "10. Optional: Windowed Vessel": "F3E5F5",
}


def style_data_row(ws, row_idx, n_cols, fill_color, border):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.border = border
        cell.font = Font(size=9.5, name="Calibri")
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_bom(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Full BOM"

    headers = ["Section", "Item", "Part Number", "Description", "Spec / Rating",
               "Vendor", "Qty", "Unit Price", "Line Total", "Direct Product Link"]
    ws.append(headers)

    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    purchasable_total = 0.0
    optional_total = 0.0
    row_idx = 2

    # ── Purchasable items (sections 1-8) ───────────────────────────────────────
    for r in ROWS:
        line_total = r["qty"] * r["unit_price"]
        purchasable_total += line_total

        ws.append([
            r["section"], r["item"], r["part"], r["desc"], r["spec"],
            r["vendor"], r["qty"], r["unit_price"], line_total,
            r.get("link_label", r["part"])
        ])
        style_data_row(ws, row_idx, len(headers), SECTION_COLORS[r["section"]], border)
        ws.cell(row=row_idx, column=8).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"

        link_cell = ws.cell(row=row_idx, column=10)
        link_cell.hyperlink = r["link"]
        link_cell.value = r.get("link_label", r["part"])
        link_cell.font = Font(size=9.5, name="Calibri", color="0563C1", underline="single")
        row_idx += 1

    # ── Subtotal: purchasable ─────────────────────────────────────────────────
    ws.append(["", "", "", "", "", "", "", "SUBTOTAL (to buy)", purchasable_total, ""])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, size=10, name="Calibri")
        cell.fill = PatternFill(start_color="D6D9E8", end_color="D6D9E8", fill_type="solid")
        cell.border = border
    ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"
    row_idx += 1

    # ── Already-owned items (section 9) ───────────────────────────────────────
    for r in ALREADY_OWNED:
        ws.append([
            r["section"], r["item"], r["part"], r["desc"], r["spec"],
            "N/A — already owned", "", "", "$0.00", "N/A"
        ])
        style_data_row(ws, row_idx, len(headers), SECTION_COLORS[r["section"]], border)
        row_idx += 1

    # ── Optional items (section 10) ───────────────────────────────────────────
    for r in OPTIONAL:
        line_total = r["qty"] * r["unit_price"]
        optional_total += line_total

        ws.append([
            r["section"], r["item"], r["part"], r["desc"], r["spec"],
            r["vendor"], r["qty"], r["unit_price"], line_total,
            r.get("link_label", r["part"])
        ])
        style_data_row(ws, row_idx, len(headers), SECTION_COLORS[r["section"]], border)
        ws.cell(row=row_idx, column=8).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"

        link_cell = ws.cell(row=row_idx, column=10)
        link_cell.hyperlink = r["link"]
        link_cell.value = r.get("link_label", r["part"])
        link_cell.font = Font(size=9.5, name="Calibri", color="0563C1", underline="single")
        row_idx += 1

    # ── Grand total row ───────────────────────────────────────────────────────
    ws.append(["", "", "", "", "", "", "",
               "TOTAL (required, excl. optional window vessel)",
               purchasable_total, ""])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, size=11, name="Calibri", color=WHITE)
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        cell.border = border
    ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"

    # Column widths
    widths = [24, 32, 20, 44, 38, 26, 6, 12, 12, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    # ── Notes sheet ────────────────────────────────────────────────────────────
    notes = wb.create_sheet("Notes & Safety")
    notes_data = [
        ["scCO2 System — Complete BOM Notes & Safety Corrections"],
        [""],
        ["LINKS: All links in this BOM are direct product page URLs named by part number."],
        ["A few items (shaft coupler, Grainger items) use search URLs where no single confirmed direct page exists."],
        [""],
        ["EXISTING HARDWARE DISCOVERED — DO NOT RE-BUY:"],
        ["  HiP valve on booster outlet → covers supply shutoff (was BV-1). Saves ~$407."],
        ["  Swagelok SS-3NRM4 on vessel outlet → covers vent isolation (was BV-3). Saves ~$407."],
        ["  Duro United 0-5000 PSI gauge on vessel → covers visual reference (was K4201). Saves ~$100."],
        ["  Result: only 1 ball valve needed (SS-83KS4 × 1 for vessel inlet isolation)."],
        [""],
        ["⚠ SS-3NRM4 MUST remain OPEN during automated operation."],
        ["  The solenoid valve (SV-1) controls automated venting. Closing the SS-3NRM4 blocks it."],
        ["  Only close SS-3NRM4 for maintenance or emergency manual isolation."],
        [""],
        ["CRITICAL SAFETY CORRECTIONS (July 2026 audit):"],
        [""],
        ["1. RELIEF VALVE — Part changed from SS-RL3S4 to SS-4R3A"],
        ["   SS-RL3S4 is Swagelok's LOW-PRESSURE series — maximum set pressure 225 PSI."],
        ["   Your system operates at 4,061 PSI. SS-RL3S4 would blow open at 225 PSI."],
        ["   Correct part: SS-4R3A (HIGH-pressure, up to 6,000 PSI)."],
        ["   Set pressure must be specified at ordering: recommend 4,350 PSI (30 MPa)."],
        [""],
        ["2. STAINLESS TUBING — Must be 0.065\" wall seamless annealed"],
        ["   Grainger 5LVR1: seamless annealed, 0.065\" wall, rated 8,125 PSI. Buy 3 × 6ft = 18ft."],
        ["   Amazon FITOK B0BB1122VJ (0.035\" wall, ~3,200 PSI) is NOT safe for this system."],
        ["   Must be SEAMLESS (not welded) for Swagelok compression fittings to seal correctly."],
        [""],
        ["3. RELAY MODULE — Must accept 3.3V RPi GPIO trigger"],
        ["   AEDIKO B095YD3732: optocoupler isolated, 1-channel, 5V coil, 3.3V-compatible input."],
        ["   GAEYAELE B07DYLKH74 requires 24VDC trigger (PLC-grade) — will not work with RPi."],
        [""],
        ["4. PRESSURE TRANSDUCER — Order item 5DEK9 (4-20 mA), NOT K4708 (1-5V DC)"],
        ["   K4708 has the wrong output type for the 250-ohm shunt wiring in this system."],
        [""],
        ["5. VENT SOLENOID VALVE — Was missing from original BOM, now added"],
        ["   Must be rated >= 6,000 PSI. Standard solenoids (150-300 PSI) will rupture at 28 MPa."],
        [""],
        ["Why exact Swagelok part numbers matter:"],
        ["- SS-CHS4-5 (check valve): 6,000 PSI. SS-4C looks similar but is only 3,000 PSI."],
        ["- SS-83KS4 (ball valve): 6,000 PSI. SS-43GS4 looks similar but is only 3,000 PSI."],
        ["- All SS-400 fittings: 5,100 PSI rated, matched to 1/4\" tube OD throughout."],
        ["- SS-1RS4 needle valve: 5,000 PSI @ 100°F — confirmed suitable for 4,061 PSI at 60°C."],
        [""],
        ["Recommended buying order:"],
        ["1. Confirm tube/thread size compatibility with PARR vessel fittings first."],
        ["2. Order Swagelok parts from an authorized distributor."],
        ["3. Specify set pressure 4,350 PSI when ordering the SS-4R3A relief valve."],
        ["4. Grainger for tubing (5LVR1) and PSU (33NT20). Amazon for electronics."],
        ["5. Source high-pressure solenoid valve from Parker or Asco distributor."],
    ]
    for row in notes_data:
        notes.append(row)
    notes["A1"].font = Font(bold=True, size=14, color=NAVY)
    notes.column_dimensions["A"].width = 115
    for r in range(3, len(notes_data) + 1):
        notes.cell(row=r, column=1).font = Font(size=10)
        notes.cell(row=r, column=1).alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    print(f"Full BOM written to: {output_path}")
    print(f"  Required purchase subtotal: ${purchasable_total:,.2f}")
    print(f"  Optional windowed vessel:   ${optional_total:,.2f}")


if __name__ == "__main__":
    build_bom("/home/user/pressure-regulation-system/scCO2_Full_Bill_of_Materials.xlsx")
