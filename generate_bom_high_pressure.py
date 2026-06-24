#!/usr/bin/env python3
"""Generate the high-pressure parts Bill of Materials as a formatted Excel file.

Links use vendor search pages (not guessed deep product URLs) so they
reliably resolve even if a vendor reorganizes their catalog.
"""

from urllib.parse import quote_plus

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY  = "1E2761"
TEAL  = "028090"
CREAM = "F8F9FA"
WHITE = "FFFFFF"
GRAY  = "6C757D"
GREEN_FILL = "E8F5E9"

def search_url(vendor, query):
    q = quote_plus(query)
    vendors = {
        "mcmaster":  f"https://www.mcmaster.com/products/?Ntt={q}",
        "amazon":    f"https://www.amazon.com/s?k={q}",
        "walmart":   f"https://www.walmart.com/search?q={q}",
        "radwell":   f"https://www.radwell.com/en-US/Search?searchTerm={q}",
        "idealvac":  f"https://www.idealvac.com/search.aspx?find={q}",
        "ebay":      f"https://www.ebay.com/sch/i.html?_nkw={q}",
        "swagelok":  f"https://www.swagelok.com/en/search?text={q}",
    }
    return vendors[vendor]

# (Item, Part Number, Description, Spec, Vendor(s), Qty, Unit Price, Search query, search vendor key)
ROWS = [
    # ── Motorized needle valve ──────────────────────────────────────────────
    dict(section="1. Motorized Needle Valve",
         item="Needle valve body",
         part="SS-1RS4",
         desc="Swagelok needle valve, regulating stem, 1/4\" tube compression",
         spec="316SS, 5,000 PSI, Cv = 0.37",
         vendor="idealvac.com / Swagelok distributor",
         qty=1, unit_price=300.00,
         search_vendor="idealvac", search_query="Swagelok SS-1RS4 needle valve"),
    dict(section="1. Motorized Needle Valve",
         item="Shaft coupler (NEMA 17 to valve stem)",
         part="5mm to 1/4\" flexible shaft coupler",
         desc="Couples existing NEMA 17 stepper directly to valve stem",
         spec="Aluminum flexible coupling, 5mm x 1/4\"",
         vendor="Amazon",
         qty=1, unit_price=10.00,
         search_vendor="amazon", search_query="flexible shaft coupler 5mm to 1/4 inch NEMA 17"),

    # ── Check valve ──────────────────────────────────────────────────────────
    dict(section="2. Check Valve",
         item="Check valve",
         part="SS-CHS4-5",
         desc="Swagelok check valve, 1/4\" tube fitting, 5 PSI cracking pressure",
         spec="316SS, 6,000 PSI (NOT SS-4C — that is only 3,000 PSI)",
         vendor="Amazon / Radwell.com",
         qty=1, unit_price=100.00,
         search_vendor="amazon", search_query="Swagelok SS-CHS4-5 check valve"),

    # ── Manual ball valves ───────────────────────────────────────────────────
    dict(section="3. Manual Ball Valves",
         item="Ball valve",
         part="SS-83KS4",
         desc="Swagelok 83 series ball valve, 1/4\" tube fitting, PCTFE seats, CO2 compatible",
         spec="316SS, 6,000 PSI (NOT SS-43GS4 — that is only 3,000 PSI)",
         vendor="Radwell.com / eBay (surplus)",
         qty=3, unit_price=407.00,
         search_vendor="radwell", search_query="Swagelok SS-83KS4 ball valve"),

    # ── Tubing ────────────────────────────────────────────────────────────────
    dict(section="4. Stainless Tubing",
         item="SS tubing, 1/4\" OD x 0.065\" wall",
         part="SS-T4-S-065-20",
         desc="Swagelok seamless 316SS tubing, 20 ft minimum order",
         spec="316SS seamless, 1/4\" OD x 0.065\" wall, 20 ft",
         vendor="mcmaster.com",
         qty=1, unit_price=120.00,
         search_vendor="mcmaster", search_query="316SS tubing 1/4 OD 0.065 wall"),

    # ── Compression fittings ─────────────────────────────────────────────────
    dict(section="5. Compression Fittings",
         item="Union (straight)",
         part="SS-400-6",
         desc="Swagelok union fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Amazon / Walmart",
         qty=4, unit_price=15.00,
         search_vendor="amazon", search_query="Swagelok SS-400-6 union fitting"),
    dict(section="5. Compression Fittings",
         item="90 degree elbow",
         part="SS-400-9",
         desc="Swagelok 90 degree elbow fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Amazon / Walmart",
         qty=3, unit_price=20.00,
         search_vendor="amazon", search_query="Swagelok SS-400-9 elbow fitting"),
    dict(section="5. Compression Fittings",
         item="Tee",
         part="SS-400-3",
         desc="Swagelok tee fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Amazon / Walmart",
         qty=2, unit_price=28.00,
         search_vendor="amazon", search_query="Swagelok SS-400-3 tee fitting"),
    dict(section="5. Compression Fittings",
         item="End cap",
         part="SS-400-C",
         desc="Swagelok end cap fitting, 1/4\" tube OD",
         spec="316SS, 5,100 PSI rated",
         vendor="Amazon",
         qty=3, unit_price=10.00,
         search_vendor="amazon", search_query="Swagelok SS-400-C end cap fitting"),
]

SECTION_COLORS = {
    "1. Motorized Needle Valve": "E0F4F7",
    "2. Check Valve": "FFF3E0",
    "3. Manual Ball Valves": "E8EAF6",
    "4. Stainless Tubing": "F1F8E9",
    "5. Compression Fittings": "FCE4EC",
}


def build_bom(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "High Pressure BOM"

    headers = ["Section", "Item", "Part Number", "Description", "Spec / Rating",
               "Vendor", "Qty", "Unit Price", "Line Total", "Buy / Search Link"]
    ws.append(headers)

    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_cost = 0.0
    row_idx = 2
    for r in ROWS:
        link = search_url(r["search_vendor"], r["search_query"])
        line_total = r["qty"] * r["unit_price"]
        total_cost += line_total

        values = [
            r["section"], r["item"], r["part"], r["desc"], r["spec"],
            r["vendor"], r["qty"], r["unit_price"], line_total, "Search this part"
        ]
        ws.append(values)

        fill_color = SECTION_COLORS.get(r["section"], "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=9.5, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.cell(row=row_idx, column=8).number_format = "$#,##0.00"
        ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"

        link_cell = ws.cell(row=row_idx, column=10)
        link_cell.hyperlink = link
        link_cell.value = "Search this part"
        link_cell.font = Font(size=9.5, name="Calibri", color="0563C1", underline="single")

        row_idx += 1

    # Total row
    ws.append(["", "", "", "", "", "", "", "TOTAL", total_cost, ""])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, size=10, name="Calibri")
        cell.fill = PatternFill(start_color="D6D9E8", end_color="D6D9E8", fill_type="solid")
        cell.border = border
    ws.cell(row=row_idx, column=9).number_format = "$#,##0.00"

    # Column widths
    widths = [22, 26, 18, 38, 32, 24, 6, 12, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    # ── Notes sheet ───────────────────────────────────────────────────────────
    notes = wb.create_sheet("Notes")
    notes_data = [
        ["High-Pressure Parts BOM — Notes"],
        [""],
        ["Why these part numbers matter:"],
        ["• SS-CHS4-5 (check valve) is rated 6,000 PSI. SS-4C looks similar but is only 3,000 PSI — not enough margin for a 4,061 PSI (28 MPa) system."],
        ["• SS-83KS4 (ball valve) is rated 6,000 PSI. SS-43GS4 looks similar but is only 3,000 PSI — same issue."],
        ["• All fittings are Swagelok SS-400 series, rated 5,100 PSI, matched to 1/4\" tube OD throughout."],
        ["• SS-1RS4 needle valve (5,000 PSI) gives ~23% margin above the 4,061 PSI operating pressure."],
        [""],
        ["Why links go to vendor search pages, not exact product URLs:"],
        ["Vendor catalog URLs change frequently and deep links can go stale or 404."],
        ["Each link searches the vendor's site for the exact part number / description so you always land on the current listing."],
        [""],
        ["Recommended buying order:"],
        ["1. Confirm thread/tube size compatibility with your existing PARR vessel fittings before ordering."],
        ["2. Order from Swagelok's official distributor network when possible for warranty/traceability on a 28 MPa system."],
        ["3. McMaster-Carr ships fastest for tubing; Swagelok parts are often cheaper through authorized distributors than Amazon resellers."],
    ]
    for row in notes_data:
        notes.append(row)
    notes["A1"].font = Font(bold=True, size=14, color=NAVY)
    notes.column_dimensions["A"].width = 110
    for r in range(3, len(notes_data) + 1):
        notes.cell(row=r, column=1).font = Font(size=10)
        notes.cell(row=r, column=1).alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    print(f"BOM written to: {output_path}  (Total: ${total_cost:,.2f})")


if __name__ == "__main__":
    build_bom("/home/user/pressure-regulation-system/scCO2_High_Pressure_BOM.xlsx")
