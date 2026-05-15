"""generate_bom.py — Generate Bill of Materials Excel sheet for scCO2 system."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Bill of Materials"

# ── Color palette ─────────────────────────────────────────────────────────────
NAVY   = "1E2761"
TEAL   = "028090"
AMBER  = "FFC107"
CREAM  = "F8F9FA"
WHITE  = "FFFFFF"
LGRAY  = "E9ECEF"
MGRAY  = "DEE2E6"

def fill(hex): return PatternFill("solid", fgColor=hex)
def font(hex, bold=False, sz=11): return Font(color=hex, bold=bold, size=sz)
def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Title row ─────────────────────────────────────────────────────────────────
ws.merge_cells("A1:I1")
ws["A1"] = "scCO₂ Pressure Regulation System — Bill of Materials"
ws["A1"].fill      = fill(NAVY)
ws["A1"].font      = Font(color=WHITE, bold=True, size=14)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:I2")
ws["A2"] = "University Research Project  |  Max Operating Pressure: 28 MPa (4,061 PSI)  |  Temperature: 60°C"
ws["A2"].fill      = fill(TEAL)
ws["A2"].font      = Font(color=WHITE, size=10)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 20

# ── Header row ────────────────────────────────────────────────────────────────
headers = ["#", "Category", "Item", "Brand", "Model / Part #",
           "Spec / Notes", "Est. Price (USD)", "Purchase Link", "Status"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.fill      = fill(MGRAY)
    c.font      = Font(bold=True, size=10, color=NAVY)
    c.alignment = center()
    c.border    = border()
ws.row_dimensions[3].height = 24

# ── Data ──────────────────────────────────────────────────────────────────────
SECTION_HEADERS = {
    "Pressure Vessel":         (TEAL,  WHITE),
    "Grainger":                (NAVY,  WHITE),
    "Amazon / Electronics":    (AMBER, NAVY),
    "High Pressure Hardware":  ("5C4033", WHITE),
}

rows = [
    # (category, item, brand, model, spec, price)
    ("Pressure Vessel",
     "Pressure vessel with sapphire window",
     "PARR Instrument",
     "Series 4790 HP",
     "100 mL, 316 SS, sapphire window, 5000 PSI, 1/4\" NPT",
     "~$2,000+ (quote)"),

    ("Grainger",
     "Pressure transducer 0–5000 PSI",
     "Ashcroft",
     "K4708",
     "1/4\" MNPT, 4–20 mA, cable leads, IP67, ±0.5%, 316 SS",
     "~$182"),

    ("Grainger",
     "Pressure gauge 0–5000 PSI",
     "Ashcroft",
     "K4201",
     "1/4\" MNPT, SS Bourdon tube, glycerin filled",
     "~$100"),

    ("Grainger",
     "Relief valve",
     "Parker",
     "442F42",
     "SS316, 1/4\" MNPT, set ~4500 PSI",
     "~$165"),

    ("Grainger",
     "DIN rail PSU 24VDC 50W",
     "Dayton",
     "33NT20",
     "24VDC output, 50W, DIN rail mount",
     "~$55"),

    ("Grainger",
     "DIN rail relay",
     "Finder",
     "Series 55",
     "24VDC coil, SPST, DIN rail mount, 10A contacts",
     "~$20"),

    ("Grainger",
     "PTFE tape (high-pressure grade)",
     "Henkel Loctite",
     "34P209",
     "1/2\" wide, high-density PTFE, for NPT threads",
     "~$5"),

    ("Amazon / Electronics",
     "Single board computer",
     "Raspberry Pi",
     "Pi 4 Model B — 4GB RAM",
     "4GB RAM, runs Python 3.11, GPIO pins for stepper + relay",
     "~$55"),

    ("Amazon / Electronics",
     "16-bit I²C ADC module",
     "Adafruit",
     "ADS1115 breakout board",
     "4-channel, 16-bit, I²C — reads pressure transducer + thermocouple",
     "~$10"),

    ("Amazon / Electronics",
     "Stepper motor",
     "Stepperonline",
     "NEMA 17, 200 steps/rev",
     "200 steps/rev, ≥40 N·cm holding torque, for needle valve actuator",
     "~$15"),

    ("Amazon / Electronics",
     "Stepper driver",
     "Pololu",
     "A4988 carrier board",
     "1/16 microstepping, up to 2A, fits NEMA 17",
     "~$8"),

    ("Amazon / Electronics",
     "microSD card",
     "Samsung",
     "32GB Pro Endurance",
     "32GB, Class 10, high-endurance for continuous logging",
     "~$10"),

    ("Amazon / Electronics",
     "Solid state relay module",
     "Opto 22",
     "DC200D",
     "24VDC control, 200VDC load, for solenoid switching",
     "~$20"),

    ("High Pressure Hardware",
     "Motorized needle valve",
     "Autoclave Engineers (Parker)",
     "SS316, stepper actuated",
     "SS316, 1/4\" tube OD compression, >5000 PSI, NEMA 17 stepper mount",
     "~$400 (quote)"),

    ("High Pressure Hardware",
     "Check valve",
     "Swagelok",
     "SS-4C-1/3",
     "SS316, 1/4\" tube OD, >5000 PSI, ~5 PSI cracking pressure",
     "~$150"),

    ("High Pressure Hardware",
     "Manual ball valve (×3)",
     "Swagelok",
     "SS-43S4",
     "SS316, 1/4\" tube OD, 6000 PSI rated — order qty 3",
     "~$100 each / $300 total"),

    ("High Pressure Hardware",
     "SS tubing 10 ft",
     "Swagelok",
     "SS-T4-S-065",
     "316SS, 1/4\" OD × 0.065\" wall, seamless, 10 ft",
     "~$80"),

    ("High Pressure Hardware",
     "Compression fittings — assorted",
     "Swagelok",
     "SS, 1/4\" tube OD",
     "Elbows, tees, unions, end caps — 316 SS, 1/4\" tube OD",
     "~$200 (assortment)"),
]

current_section = None
row_num = 4
item_num = 1

for (category, item, brand, model, spec, price) in rows:
    # Section header
    if category != current_section:
        current_section = category
        bg, fg = SECTION_HEADERS[category]
        ws.merge_cells(f"A{row_num}:I{row_num}")
        c = ws.cell(row=row_num, column=1, value=f"  {category.upper()}")
        c.fill      = fill(bg)
        c.font      = Font(color=fg, bold=True, size=10)
        c.alignment = left()
        c.border    = border()
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # Data row
    bg_color = WHITE if item_num % 2 == 1 else CREAM
    values = [item_num, category, item, brand, model, spec, price, "", "☐ To Order"]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill      = fill(bg_color)
        c.font      = font("333333")
        c.border    = border()
        c.alignment = center() if col in (1, 7, 9) else left()
    ws.row_dimensions[row_num].height = 36
    row_num += 1
    item_num += 1

# ── Total row ─────────────────────────────────────────────────────────────────
ws.merge_cells(f"A{row_num}:F{row_num}")
c = ws.cell(row=row_num, column=1, value="ESTIMATED TOTAL (excluding vessel & high-pressure hardware quotes)")
c.fill = fill(NAVY); c.font = Font(color=WHITE, bold=True, size=10)
c.alignment = left(); c.border = border()

c = ws.cell(row=row_num, column=7, value="~$1,820 + quotes")
c.fill = fill(NAVY); c.font = Font(color=WHITE, bold=True, size=11)
c.alignment = center(); c.border = border()

for col in (8, 9):
    c = ws.cell(row=row_num, column=col)
    c.fill = fill(NAVY); c.border = border()

ws.row_dimensions[row_num].height = 24

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [4, 22, 36, 22, 24, 52, 20, 30, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze header rows ────────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Notes sheet ──────────────────────────────────────────────────────────────
ns = wb.create_sheet("Notes & Contacts")
notes = [
    ("Vendor", "Contact", "Website", "Notes"),
    ("PARR Instrument", "(309) 762-7716", "parrinst.com", "Call for sapphire window quote on 4790 HP"),
    ("Swagelok", "Local distributor", "swagelok.com", "Find nearest distributor on their website"),
    ("Autoclave Engineers", "Parker Hannifin", "autoclaveengineers.com", "Request quote for motorized needle valve with stepper actuator"),
    ("Grainger", "Online / local branch", "grainger.com", "Use part numbers listed in BOM tab"),
    ("Adafruit", "Online only", "adafruit.com", "ADS1115 breakout — also on Amazon"),
    ("Pololu", "Online only", "pololu.com", "A4988 stepper driver"),
    ("Stepperonline", "Online only", "omc-stepperonline.com", "NEMA 17 stepper motor"),
]
for r, row in enumerate(notes, 1):
    for c, val in enumerate(row, 1):
        cell = ns.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = fill(NAVY)
        cell.alignment = left()
        cell.border = border()

for i, w in enumerate([20, 20, 25, 50], 1):
    ns.column_dimensions[get_column_letter(i)].width = w

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/home/user/pressure-regulation-system/scCO2_Bill_of_Materials.xlsx"
wb.save(out)
print(f"Saved: {out}")
